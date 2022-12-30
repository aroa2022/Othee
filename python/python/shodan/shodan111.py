import paramiko
from shodan import Shodan
import openpyxl
from openpyxl import Workbook,load_workbook


lled= 'IdTEUOuT6iumTxmuLVS6q28w5pnPxfoS'
a3='H1a0bvHBOroQop2r2hJ8fk1nr8CK3dVV'

def f(ip):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname=ip, username="admin", password="admin")
        stdin, stdout, stderr = ssh.exec_command("ip firewall connection print")
        print(stdout.read().decode("ascii").strip("\n"))
        ssh.close()
    except:
        print('error')

wb = load_workbook('mikrotik.xlsx')
sheet = wb.active

api = Shodan('H1a0bvHBOroQop2r2hJ8fk1nr8CK3dVV')
results = api.search('Dahua DVR')
data = results['matches']
for i in data:
    c= sheet.max_row
    print(i['ip_str'],':',i['port'])
    sheet[f'A{c+1}'] =i['ip_str']
    sheet[f'B{c+1}'] =str(i['port'])

wb.save('mikrotik.xlsx')
