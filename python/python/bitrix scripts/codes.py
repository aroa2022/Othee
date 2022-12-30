from selenium import  webdriver
import pickle
web = webdriver.Chrome()

###selenium ################
# export coockie from browser
pickle.dump( web.get_cookies() , open("cookies.pkl","wb"))
# import  coockie to browser
cookies = pickle.load(open("cookies.pkl", "rb"))
for cookie in cookies:
    web.add_cookie(cookie)

###############################################openpyxl#########################################
import openpyxl
wb = openpyxl.Workbook
sheet = wb.create_sheet()
wb =  openpyxl.Workbook()
sheet['A1'] = 'name'
sheet['B1'] = 'last name'
sheet = wb.create_sheet()
wb.save('db.xlsx')

