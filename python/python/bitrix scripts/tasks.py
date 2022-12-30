from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
from time import sleep
from pprint import pprint
get_task = 'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/task.item.list.json?start=100'
get_emp = 'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.get.json?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46'
update_emp='https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID='
def task_info():
    r= requests.get(get_task).json()
    for task in r['result']:
        title = task['TITLE']
        description = task['DESCRIPTION']
        DEADLINE = task['DEADLINE']
        START_DATE_PLAN = task['START_DATE_PLAN']
        START_DATE_PLAN = task['START_DATE_PLAN']
        GROUP_ID = task['GROUP_ID']
        RESPONSIBLE_ID = task['RESPONSIBLE_ID']
        T_ID =task['ID']
        creeated_by = task['CREATED_BY']
        RESPONSIBLE_NAME = task['RESPONSIBLE_NAME']
        CREATED_BY_LAST_NAME = task['CREATED_BY_LAST_NAME']
        CREATED_DATE =task['CREATED_DATE']
        status = task['REAL_STATUS']
        CLOSED_BY=task['CLOSED_BY']

        print(status)


def emp_info():
    workbook1 = load_workbook(filename="db.xlsx")
    sheet1 = workbook1.active
    task_in_month2=[]
    project2=[]
    user = []

    for i in range(1,int(int(sheet1.max_row)+1)):
        sp=sheet1[f'A{i}'].value
        user .append(sp.split(':')[0])
        task_in_month2.append(sp.split(':')[1])
        project2.append(sp.split(':')[2])




    c2=0
    c= 0 
    r = requests.get(get_emp).json()
    for emp in r['result']:
        
        
        emp_name=emp['NAME']
        emp_id=emp['ID']
        #task_in_month
        for k ,l in emp['PROPERTY_600'].items():
            emp_task_in_month = l 
        #task in project
        for k ,l in emp['PROPERTY_602'].items():
            emp_project_in_month = l
        #reward
        for k ,l in emp['PROPERTY_606'].items():
            emp_reward = l
        #salary
        for k ,l in emp['PROPERTY_604'].items():
            emp_salary = l
        #print(emp_name,emp_project_in_month,emp_task_in_month) 
        if int(emp_task_in_month) >118 and int(emp_task_in_month) < 149 and int(emp_project_in_month)>4:
            r= requests.get(f'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID={emp_id}&FIELDS[NAME]={emp_name}&FIELDS[PROPERTY_600]={emp_task_in_month}&FIELDS[PROPERTY_602]={emp_project_in_month}&FIELDS[PROPERTY_604]={emp_salary}&FIELDS[PROPERTY_606]={int(emp_salary)*5/100}')
        if int(emp_task_in_month) >149 and int(emp_project_in_month)>4:
            r= requests.get(f'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID={emp_id}&FIELDS[NAME]={emp_name}&FIELDS[PROPERTY_600]={emp_task_in_month}&FIELDS[PROPERTY_602]={emp_project_in_month}&FIELDS[PROPERTY_604]={emp_salary}&FIELDS[PROPERTY_606]={int(emp_salary)*10/100}')
                                         
        
        r= requests.get(get_task).json()
        for task in r['result']:
            c2=c2+1
            print(c2)
            sleep(5)
            status = task['REAL_STATUS']
            RESPONSIBLE_NAME = task['RESPONSIBLE_NAME']
            T_ID =task['ID']
            title = task['TITLE']
            group_id = task['GROUP_ID']
            print('task id ' ,T_ID,'*'*30)
            if emp_name in RESPONSIBLE_NAME :
                
                
                if status == '5':
                    if T_ID in task_in_month2:
                        print('task in database')

                    if T_ID not in task_in_month2 and int(group_id) ==0:
                        print('new task finish with out project')
                        r= requests.get(f'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID={emp_id}&FIELDS[NAME]={emp_name}&FIELDS[PROPERTY_600]={int(emp_task_in_month)+1}&FIELDS[PROPERTY_602]={emp_project_in_month}&FIELDS[PROPERTY_604]={emp_salary}&FIELDS[PROPERTY_606]={emp_reward}')     
                        sheet1[f"A{(sheet1.max_row+1)}"] = f"{RESPONSIBLE_NAME}:{T_ID}:0"
                        workbook1.save(filename="db.xlsx")

                    if T_ID not in task_in_month2 and  int(group_id) >0 and group_id in project2:
                        print('new task finish and project in database')
                        r= requests.get(f'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID={emp_id}&FIELDS[NAME]={emp_name}&FIELDS[PROPERTY_600]={int(emp_task_in_month)+1}&FIELDS[PROPERTY_602]={emp_project_in_month}&FIELDS[PROPERTY_604]={emp_salary}&FIELDS[PROPERTY_606]={emp_reward}')
                        sheet1[f"A{(sheet1.max_row+1)}"] = f"{RESPONSIBLE_NAME}:{T_ID}:{group_id}"
                        workbook1.save(filename="db.xlsx")
                    if T_ID not in task_in_month2 and int(group_id)>0 and group_id not in project2:
                            print('new task finished and group id not in database')
                            r= requests.get(f'https://b24-vgcuh0.bitrix24.com/rest/10/dhx2b3zg2lq30l1p/lists.element.update?IBLOCK_TYPE_ID=bitrix_processes&IBLOCK_ID=46&ELEMENT_ID={emp_id}&FIELDS[NAME]={emp_name}&FIELDS[PROPERTY_600]={int(emp_task_in_month)+1}&FIELDS[PROPERTY_602]={int(emp_project_in_month)+1}&FIELDS[PROPERTY_604]={emp_salary}&FIELDS[PROPERTY_606]={emp_reward}')
                            
                            sheet1[f"A{(sheet1.max_row+1)}"] =f"{RESPONSIBLE_NAME}:{T_ID}:{group_id}"
                            workbook1.save(filename="db.xlsx")
                            
        
   




emp_info()


